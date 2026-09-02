"""QA do Track One — o app novo que os técnicos vão usar do chamado até o
fechamento da RAT.

Cada técnico (ou líder de equipe) convidado a testar vira uma linha aqui, com
o funil de QA (a contatar -> convidado -> instalado -> em teste -> concluído)
e o histórico do que ele relatou no teste (o que achou bom, o que precisa
melhorar, o que deu problema). Este módulo também monta a mensagem de convite
pronta pra abrir no WhatsApp — o texto muda conforme o papel (técnico direto
ou líder avisando a equipe), com o nome já substituído.
"""
import io
import unicodedata
from datetime import datetime, timedelta, timezone
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session, joinedload

from .. import formulario, models, schemas
from ..database import get_db
from . import ativos

router = APIRouter(tags=["tecnicos"])
# páginas abertas no navegador (sem o prefixo /api) — o formulário que o técnico
# recebe por WhatsApp e abre no celular
pagina_router = APIRouter(tags=["tecnicos"])

BRT = timezone(timedelta(hours=-3))

PAPEIS = {"tecnico", "lider"}
# funil de QA: do convite até o teste no atendimento real virar feedback registrado
STATUSES = ["a_contatar", "convidado", "instalado", "em_teste", "concluido", "sem_retorno"]
STATUS_LABELS = {
    "a_contatar": "A contatar",
    "convidado": "Convidado",
    "instalado": "App instalado",
    "em_teste": "Em teste",
    "concluido": "Teste concluído",
    "sem_retorno": "Sem retorno",
}
TIPOS_OBS = {"positivo", "melhoria", "problema"}

# textos enviados exatamente como definidos com o Rafa — {nome} é a única
# substituição feita na hora de gerar o link do WhatsApp.
TEMPLATE_TECNICO = """Fala, {nome}! Tudo certo?
Estamos lançando um app novo pra técnicos (Track One) e você foi selecionado pra testar antes de liberar geral.
O que ele faz:

* Acompanha o fluxo inteiro do atendimento, desde o chamado atribuído a você até o fechamento da RAT, tudo pelo app
* Mostra rastreio e previsão de entrega quando o atendimento precisa de peça
* Você confirma o recebimento do equipamento direto por lá

Vou te chamar pra fazer a instalação e já passo o manual de uso na hora. Depois é só usar normal no seu próximo atendimento, do começo ao fim, e qualquer coisa estranha (tela que não atualiza, notificação que não chega, informação que falta) me avisa direto — print ajuda muito.
Bora marcar a instalação?"""

TEMPLATE_LIDER = """Fala, {nome}! Tudo certo?
Estamos lançando um app novo pra técnicos (Track One) e já vou entrar em contato direto com o seu time pra fazer a instalação. Só queria te avisar antes.
O que ele faz:

* Acompanha o fluxo inteiro do atendimento, desde o chamado atribuído ao técnico até o fechamento da RAT, tudo pelo app
* Mostra rastreio e previsão de entrega quando precisa de peça
* O técnico confirma o recebimento do equipamento direto por lá

Tem manual de uso, vou passar junto na instalação com cada um. Pode avisar o pessoal que eu vou chamar eles nos próximos dias pra instalar e usar no próximo atendimento?"""

# mandada DEPOIS do atendimento, com o link do formulário — é o que fecha o ciclo
# do teste sem precisar entrevistar cada técnico por mensagem
TEMPLATE_FEEDBACK = """Fala, {nome}! Tudo certo?
Vi que você usou o Track One no atendimento — me conta rapidinho como foi?
São 2 minutinhos, direto no link:

{link}

Pode ser sincero, é justamente pra ajustar o que estiver ruim antes de liberar pra todo mundo. Valeu demais!"""

# cobrança de quem parou no meio do caminho — muda conforme onde ele travou, que
# é a diferença entre "instala aí" e "usa no próximo atendimento"
TEMPLATES_COBRANCA = {
    "convidado": """Fala, {nome}! Tudo certo?
Passando pra lembrar do Track One — ficou de marcar a instalação comigo e ainda não conseguimos fechar.
São 10 minutinhos pra instalar e eu te passo o manual na hora. Quando fica bom pra você?""",
    "instalado": """Fala, {nome}! Tudo certo?
Você já está com o Track One instalado — é só usar no seu próximo atendimento, do começo ao fim.
Qualquer coisa estranha me chama, e no fim eu te mando um link rapidinho pra contar como foi. Fechou?""",
    "em_teste": """Fala, {nome}! Tudo certo?
Vi que você já usou o Track One no atendimento — só falta me contar como foi, são 2 minutinhos:

{link}

Seu retorno é o que ajusta o app antes de liberar pra todo mundo. Valeu!""",
}


def _norm_txt(valor) -> str:
    """Normaliza texto de célula de planilha pra comparar sem acento/maiúscula:
    "Líder" e "lider" caem no mesmo lugar, célula vazia (None) vira string vazia."""
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c))


def _norm_telefone(valor: str) -> str:
    digitos = "".join(c for c in (valor or "") if c.isdigit())
    if not digitos:
        raise HTTPException(status_code=400, detail="Telefone vazio.")
    # sem DDI (10 ou 11 dígitos = DDD + número) -> assume Brasil
    if len(digitos) in (10, 11):
        digitos = "55" + digitos
    if len(digitos) < 12:
        raise HTTPException(status_code=400, detail="Telefone inválido — inclua DDD (e DDI, se não for Brasil).")
    return digitos


def _mensagem_para(tecnico: models.Tecnico, tipo: str = "convite", base_url: str = "") -> str:
    """Monta o texto pronto pra mandar: o convite (antes do teste, conforme o
    papel), o pedido de feedback com o link do formulário (depois dele) ou a
    cobrança de quem travou no meio (que muda conforme a etapa onde parou)."""
    primeiro_nome = (tecnico.nome or "").strip().split(" ")[0] or tecnico.nome
    link = f"{(base_url or '').rstrip('/')}/formulario/{tecnico.token or ''}"
    if tipo == "cobranca":
        # quem ainda nem foi convidado não tem o que cobrar: recebe o convite
        template = TEMPLATES_COBRANCA.get(tecnico.status)
        if not template:
            return _mensagem_para(tecnico, "convite", base_url)
        return template.format(nome=primeiro_nome, link=link)
    if tipo == "feedback":
        return TEMPLATE_FEEDBACK.format(nome=primeiro_nome, link=link)
    template = TEMPLATE_LIDER if tecnico.papel == "lider" else TEMPLATE_TECNICO
    return template.format(nome=primeiro_nome)


def _versao_da_fase(tecnico: models.Tecnico) -> Optional[str]:
    """A build que a fase do técnico está testando agora. Gravada em cada relato
    no momento em que ele entra: se a LP subir uma versão nova no meio do piloto,
    o que já foi relatado continua apontando pra build em que aconteceu."""
    return tecnico.fase.versao_app if tecnico.fase else None


def _get_tecnico_or_404(db: Session, tecnico_id: int) -> models.Tecnico:
    tecnico = db.query(models.Tecnico).filter(models.Tecnico.id == tecnico_id).first()
    if not tecnico:
        raise HTTPException(status_code=404, detail="Técnico não encontrado")
    return tecnico


FASE_STATUSES = ["planejada", "em_andamento", "concluida", "liberada"]


@router.get("/tecnicos", response_model=List[schemas.TecnicoOut])
def list_tecnicos(
    status: Optional[str] = None,
    papel: Optional[str] = None,
    regional: Optional[str] = None,
    busca: Optional[str] = None,
    fase_id: Optional[int] = None,
    limite: int = 300,
    db: Session = Depends(get_db),
):
    """Lista completa (com observações) — é o que alimenta os cards.

    Sempre limitada: a base tem milhares de cadastros e a tela trabalha uma fase
    por vez. Pra varrer a base inteira, use /tecnicos/base, que é leve e paginada.
    """
    q = (
        db.query(models.Tecnico)
        .options(joinedload(models.Tecnico.observacoes).joinedload(models.TecnicoObservacao.ajuste))
    )
    if fase_id is not None:
        q = q.filter(models.Tecnico.fase_id == (fase_id if fase_id > 0 else None))
    if status:
        q = q.filter(models.Tecnico.status == status)
    if papel:
        q = q.filter(models.Tecnico.papel == papel)
    if regional:
        q = q.filter(models.Tecnico.regional == regional)
    if busca:
        termo = f"%{busca.strip()}%"
        q = q.filter(models.Tecnico.nome.ilike(termo))
    return q.order_by(models.Tecnico.nome).limit(max(1, min(limite, 1000))).all()


@router.get("/tecnicos/base", response_model=schemas.BaseTecnicosOut)
def base_tecnicos(
    busca: Optional[str] = None,
    regional: Optional[str] = None,
    papel: Optional[str] = None,
    sem_fase: bool = False,
    limite: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """A base inteira em formato leve e paginado — sem observações, sem token.

    É o que a tela usa pra escolher quem entra numa fase: com quatro mil
    cadastros, mandar a lista completa era mais de um mega de JSON por abertura.
    """
    q = db.query(models.Tecnico)
    if sem_fase:
        q = q.filter(models.Tecnico.fase_id.is_(None))
    if regional:
        q = q.filter(models.Tecnico.regional == regional)
    if papel:
        q = q.filter(models.Tecnico.papel == papel)
    if busca:
        q = q.filter(models.Tecnico.nome.ilike(f"%{busca.strip()}%"))
    total = q.count()
    itens = (
        q.order_by(models.Tecnico.nome)
        .offset(max(0, offset))
        .limit(max(1, min(limite, 200)))
        .all()
    )
    return {"total": total, "itens": itens}


@router.get("/tecnicos/regionais")
def listar_regionais(db: Session = Depends(get_db)):
    """As regionais que existem na base, com quantos técnicos cada uma tem — é
    por aqui que se monta uma fase por região sem digitar o nome na mão."""
    linhas = (
        db.query(models.Tecnico.regional, sa_func.count(models.Tecnico.id))
        .filter(models.Tecnico.regional.isnot(None), models.Tecnico.regional != "")
        .group_by(models.Tecnico.regional)
        .order_by(sa_func.count(models.Tecnico.id).desc())
        .all()
    )
    return [{"regional": r, "total": n} for r, n in linhas]


@router.get("/piloto/fases", response_model=List[schemas.PilotoFaseOut])
def listar_fases(db: Session = Depends(get_db)):
    fases = db.query(models.PilotoFase).order_by(
        models.PilotoFase.ordem, models.PilotoFase.id
    ).all()
    contagem = dict(
        db.query(models.Tecnico.fase_id, sa_func.count(models.Tecnico.id))
        .filter(models.Tecnico.fase_id.isnot(None))
        .group_by(models.Tecnico.fase_id)
        .all()
    )
    saida = []
    for f in fases:
        dados = schemas.PilotoFaseOut.model_validate(f)
        dados.total_tecnicos = contagem.get(f.id, 0)
        saida.append(dados)
    return saida


@router.post("/piloto/fases", response_model=schemas.PilotoFaseOut, status_code=201)
def criar_fase(payload: schemas.PilotoFaseCreate, db: Session = Depends(get_db)):
    nome = (payload.nome or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome da fase vazio.")
    maior = db.query(sa_func.max(models.PilotoFase.ordem)).scalar() or 0
    fase = models.PilotoFase(
        nome=nome,
        descricao=payload.descricao or "",
        status="planejada",
        ordem=maior + 1,
        versao_app=(payload.versao_app or "").strip() or None,
        meta_concluidos=payload.meta_concluidos,
        meta_nota=payload.meta_nota,
        meta_etapa=payload.meta_etapa,
        autor=(payload.autor or "").strip() or None,
    )
    db.add(fase)
    db.commit()
    db.refresh(fase)
    return fase


@router.patch("/piloto/fases/{fase_id}", response_model=schemas.PilotoFaseOut)
def atualizar_fase(fase_id: int, payload: schemas.PilotoFaseUpdate, db: Session = Depends(get_db)):
    fase = db.query(models.PilotoFase).filter(models.PilotoFase.id == fase_id).first()
    if not fase:
        raise HTTPException(status_code=404, detail="Fase não encontrada")
    if payload.nome is not None:
        nome = payload.nome.strip()
        if not nome:
            raise HTTPException(status_code=400, detail="Nome da fase vazio.")
        fase.nome = nome
    if payload.descricao is not None:
        fase.descricao = payload.descricao
    if payload.status is not None:
        if payload.status not in FASE_STATUSES:
            raise HTTPException(status_code=400, detail="Situação de fase inválida.")
        fase.status = payload.status
        agora = datetime.now(timezone.utc)
        if payload.status == "em_andamento" and fase.iniciada_em is None:
            fase.iniciada_em = agora
        if payload.status == "liberada" and fase.liberada_em is None:
            fase.liberada_em = agora
    if payload.ordem is not None:
        fase.ordem = payload.ordem
    if payload.versao_app is not None:
        fase.versao_app = payload.versao_app.strip() or None
    for campo in ("meta_concluidos", "meta_nota", "meta_etapa"):
        valor = getattr(payload, campo)
        if valor is not None:
            setattr(fase, campo, valor or None)
    db.commit()
    db.refresh(fase)
    return fase


@router.delete("/piloto/fases/{fase_id}")
def excluir_fase(fase_id: int, db: Session = Depends(get_db)):
    """Remove a fase. Os técnicos não são apagados — voltam pra base geral."""
    fase = db.query(models.PilotoFase).filter(models.PilotoFase.id == fase_id).first()
    if not fase:
        raise HTTPException(status_code=404, detail="Fase não encontrada")
    soltos = (
        db.query(models.Tecnico).filter(models.Tecnico.fase_id == fase_id)
        .update({"fase_id": None}, synchronize_session=False)
    )
    db.delete(fase)
    db.commit()
    return {"deleted": fase_id, "tecnicos_soltos": soltos}


@router.post("/piloto/fases/{fase_id}/tecnicos")
def adicionar_na_fase(fase_id: int, payload: schemas.AdicionarNaFase, db: Session = Depends(get_db)):
    """Coloca técnicos na fase: por lista de ids ou por filtro da base.

    O filtro é o que faz "toda a regional de Campinas" entrar de uma vez. Quem já
    está em outra fase fica de fora por padrão — mover alguém de fase tem que ser
    escolha explícita, senão uma seleção ampla esvazia a fase do vizinho.
    """
    fase = db.query(models.PilotoFase).filter(models.PilotoFase.id == fase_id).first()
    if not fase:
        raise HTTPException(status_code=404, detail="Fase não encontrada")

    q = db.query(models.Tecnico)
    if payload.tecnico_ids:
        q = q.filter(models.Tecnico.id.in_(payload.tecnico_ids))
    else:
        if payload.regional:
            q = q.filter(models.Tecnico.regional == payload.regional)
        if payload.papel:
            q = q.filter(models.Tecnico.papel == payload.papel)
        if payload.busca:
            q = q.filter(models.Tecnico.nome.ilike(f"%{payload.busca.strip()}%"))
        if not (payload.regional or payload.papel or payload.busca):
            raise HTTPException(status_code=400, detail="Escolha técnicos ou informe um filtro.")
    if not payload.incluir_de_outras_fases:
        q = q.filter((models.Tecnico.fase_id.is_(None)) | (models.Tecnico.fase_id == fase_id))

    alvos = q.all()
    movidos = 0
    for t in alvos:
        if t.fase_id != fase_id:
            t.fase_id = fase_id
            movidos += 1
    # colocar gente na fase é o que a tira do papel
    if movidos and fase.status == "planejada":
        fase.status = "em_andamento"
        fase.iniciada_em = fase.iniciada_em or datetime.now(timezone.utc)
    db.commit()
    total = db.query(models.Tecnico).filter(models.Tecnico.fase_id == fase_id).count()
    return {"adicionados": movidos, "na_fase": total}


@router.delete("/piloto/fases/{fase_id}/tecnicos/{tecnico_id}")
def tirar_da_fase(fase_id: int, tecnico_id: int, db: Session = Depends(get_db)):
    tecnico = _get_tecnico_or_404(db, tecnico_id)
    if tecnico.fase_id != fase_id:
        raise HTTPException(status_code=400, detail="Esse técnico não está nessa fase.")
    tecnico.fase_id = None
    db.commit()
    return {"tecnico_id": tecnico_id, "fase_id": None}


@router.get("/tecnicos/resumo")
def resumo_tecnicos(db: Session = Depends(get_db)):
    tecnicos = db.query(models.Tecnico).options(joinedload(models.Tecnico.observacoes)).all()
    counts = {s: 0 for s in STATUSES}
    feedback = {"positivo": 0, "melhoria": 0, "problema": 0}
    for t in tecnicos:
        counts[t.status] = counts.get(t.status, 0) + 1
        for o in t.observacoes:
            if o.tipo in feedback:
                feedback[o.tipo] += 1
    return {"total": len(tecnicos), "counts": counts, "feedback": feedback}


# palavras que aparecem em quase todo relato e não dizem nada sobre o problema —
# sem tirá-las, o "mais citado" viraria "que, não, para, app"
TERMOS_IGNORADOS = {
    "para", "pelo", "pela", "como", "mais", "muito", "quando", "porque", "também",
    "ainda", "então", "isso", "esse", "essa", "aqui", "está", "estava", "tem",
    "ter", "foi", "ser", "sempre", "nunca", "tudo", "todo", "toda", "cada",
    "mas", "sem", "com", "uma", "num", "numa", "dos", "das", "nos", "nas",
    "que", "não", "sim", "vez", "vezes", "pra", "pro", "meu", "minha", "seu",
    "sua", "ele", "ela", "eles", "elas", "você", "voce", "gente", "coisa",
    # domínio: aparecem em todo relato do piloto e não separam um problema do outro
    "app", "aplicativo", "track", "one", "sistema", "celular", "consegui",
    "consegue", "fazer", "ficou", "ficar", "deu", "dar", "usar", "uso",
}


def _termos_mais_citados(observacoes, limite: int = 8):
    """Conta as palavras que mais aparecem nos relatos de problema e melhoria.

    É uma heurística de leitura rápida, não classificação: serve pra alguém bater
    o olho e perceber "notificação apareceu 12 vezes" antes de abrir relato por
    relato. O agrupamento que vale de verdade é o vínculo com o ajuste."""
    contagem = {}
    forma_original = {}
    for o in observacoes:
        if o.tipo not in ("problema", "melhoria"):
            continue
        vistos_na_nota = set()   # a mesma palavra repetida num relato conta uma vez
        for palavra in (o.texto or "").split():
            limpa = "".join(c for c in palavra if c.isalpha() or c == "-").strip("-")
            if len(limpa) < 4:
                continue
            chave = _norm_txt(limpa)
            if chave in TERMOS_IGNORADOS or chave in vistos_na_nota:
                continue
            vistos_na_nota.add(chave)
            contagem[chave] = contagem.get(chave, 0) + 1
            forma_original.setdefault(chave, limpa.lower())
    ordenados = sorted(contagem.items(), key=lambda kv: (-kv[1], kv[0]))
    return [
        {"termo": forma_original[chave], "total": n}
        for chave, n in ordenados[:limite] if n > 1   # citado uma vez só não é padrão
    ]


@router.get("/tecnicos/piloto")
def painel_piloto(
    fase_id: Optional[int] = None,
    meta_concluidos: int = 10,
    meta_nota: float = 4.0,
    meta_etapa: int = 5,
    dias_parado: int = 3,
    db: Session = Depends(get_db),
):
    """Os números que respondem "já dá pra liberar o Track One pra todo mundo?".

    Junta o funil de adoção, a cobertura de cada etapa do fluxo (de nada adianta
    40 técnicos testando se ninguém fechou RAT pelo app), as notas, os termos que
    mais aparecem nos relatos e os ajustes que já saíram do piloto — e fecha com
    os critérios de liberação batidos contra as metas.
    """
    # o piloto anda por fase, então a régua também é por fase: uma fase de 12
    # técnicos numa capital não tem a mesma meta de uma de 60 no interior
    fase = None
    if fase_id:
        fase = db.query(models.PilotoFase).filter(models.PilotoFase.id == fase_id).first()
        if not fase:
            raise HTTPException(status_code=404, detail="Fase não encontrada")
        meta_concluidos = fase.meta_concluidos or meta_concluidos
        meta_nota = fase.meta_nota or meta_nota
        meta_etapa = fase.meta_etapa or meta_etapa

    q = db.query(models.Tecnico).options(
        joinedload(models.Tecnico.observacoes).joinedload(models.TecnicoObservacao.ajuste)
    )
    if fase_id:
        q = q.filter(models.Tecnico.fase_id == fase_id)
    tecnicos = q.all()
    total = len(tecnicos)
    por_status = {s: 0 for s in STATUSES}
    for t in tecnicos:
        por_status[t.status] = por_status.get(t.status, 0) + 1

    # no funil, cada etapa conta quem chegou nela OU passou dela — é o que mostra
    # onde o piloto está travando, e não só onde cada um parou
    caminho = [s for s in STATUSES if s != "sem_retorno"]
    alcancou = {}
    for i, etapa in enumerate(caminho):
        alcancou[etapa] = sum(por_status.get(e, 0) for e in caminho[i:])
    funil = [
        {
            "key": etapa,
            "label": STATUS_LABELS[etapa],
            "total": alcancou[etapa],
            "pct": round(alcancou[etapa] / total * 100) if total else 0,
            "parados": por_status.get(etapa, 0),
        }
        for etapa in caminho
    ]

    responderam = [t for t in tecnicos if t.respondido_em]
    etapas_contagem = {etapa: 0 for etapa in formulario.ETAPAS}
    for t in tecnicos:
        for etapa in (t.etapas_testadas or "").split("|"):
            if etapa in etapas_contagem:
                etapas_contagem[etapa] += 1
    cobertura = [
        {
            "etapa": etapa,
            "curto": formulario.ETAPAS_CURTAS.get(etapa, etapa),
            "total": n,
            "pct": round(n / len(responderam) * 100) if responderam else 0,
            "ok": n >= meta_etapa,
        }
        for etapa, n in etapas_contagem.items()
    ]

    notas = [t.nota for t in tecnicos if t.nota]
    media = round(sum(notas) / len(notas), 1) if notas else None
    distribuicao = [{"nota": n, "total": sum(1 for x in notas if x == n)} for n in (1, 2, 3, 4, 5)]

    todas_obs = [o for t in tecnicos for o in t.observacoes]
    relatos = {
        "positivo": sum(1 for o in todas_obs if o.tipo == "positivo"),
        "melhoria": sum(1 for o in todas_obs if o.tipo == "melhoria"),
        "problema": sum(1 for o in todas_obs if o.tipo == "problema"),
        "no_backlog": sum(1 for o in todas_obs if o.ajuste_id),
    }

    # ajustes que nasceram do piloto, ordenados por quantos técnicos relataram o
    # mesmo ponto — é a fila de prioridade real pra levar pra LP
    por_ajuste = {}
    for o in todas_obs:
        if o.ajuste_id:
            por_ajuste.setdefault(o.ajuste_id, []).append(o)
    ranking = []
    for ajuste_id, obs_list in por_ajuste.items():
        ajuste = obs_list[0].ajuste
        if not ajuste:
            continue
        ranking.append({
            "ajuste_id": ajuste_id,
            "ref": f"{ajuste.versao} #{ajuste.numero:02d}",
            "titulo": ajuste.titulo,
            "tipo": ajuste.tipo,
            "status": ajuste.status,
            "relatos": len(obs_list),
        })
    ranking.sort(key=lambda a: (-a["relatos"], a["ref"]))

    # quem travou no meio do funil: o piloto morre de silêncio, não de bug, então
    # essa lista é o que evita perder técnico por falta de cobrança
    agora = datetime.now(timezone.utc)

    def _dias(quando):
        if not quando:
            return None
        if quando.tzinfo is None:
            quando = quando.replace(tzinfo=timezone.utc)
        return (agora - quando).days

    parados = []
    for t in tecnicos:
        if t.status not in ("convidado", "instalado", "em_teste"):
            continue
        marco = {"convidado": t.convidado_em, "instalado": t.instalado_em}.get(t.status) or t.updated_at
        dias = _dias(marco)
        if dias is None or dias < dias_parado:
            continue
        parados.append({
            "id": t.id, "nome": t.nome, "status": t.status,
            "status_label": STATUS_LABELS[t.status], "dias": dias,
        })
    parados.sort(key=lambda p: -p["dias"])

    concluidos = por_status.get("concluido", 0)
    etapas_ok = sum(1 for c in cobertura if c["ok"])
    criterios = [
        {
            "nome": f"{meta_concluidos} técnicos com teste concluído",
            "atual": concluidos, "meta": meta_concluidos, "ok": concluidos >= meta_concluidos,
        },
        {
            "nome": f"Nota média igual ou acima de {meta_nota}",
            "atual": media if media is not None else 0, "meta": meta_nota,
            "ok": media is not None and media >= meta_nota,
        },
        {
            "nome": f"Cada etapa testada por {meta_etapa}+ técnicos",
            "atual": etapas_ok, "meta": len(formulario.ETAPAS),
            "ok": etapas_ok == len(formulario.ETAPAS),
        },
        {
            "nome": "Nenhum problema relatado fora do backlog",
            "atual": relatos["no_backlog"], "meta": relatos["problema"],
            "ok": relatos["problema"] == 0 or relatos["no_backlog"] >= relatos["problema"],
        },
    ]

    return {
        "fase": (
            {"id": fase.id, "nome": fase.nome, "status": fase.status,
             "liberada_em": fase.liberada_em.isoformat() if fase.liberada_em else None}
            if fase else None
        ),
        "total": total,
        "responderam": len(responderam),
        "funil": funil,
        "sem_retorno": por_status.get("sem_retorno", 0),
        "cobertura": cobertura,
        "notas": {"media": media, "respostas": len(notas), "distribuicao": distribuicao},
        "relatos": relatos,
        "termos": _termos_mais_citados(todas_obs),
        "parados": parados[:8],
        "parados_total": len(parados),
        "dias_parado": dias_parado,
        "ranking_ajustes": ranking[:6],
        "criterios": criterios,
        "liberado": all(c["ok"] for c in criterios),
    }


@router.post("/tecnicos/observacoes/{observacao_id}/vincular-ajuste", response_model=schemas.TecnicoOut)
def vincular_ajuste(observacao_id: int, payload: schemas.VincularAjuste, db: Session = Depends(get_db)):
    """Aponta o relato para um ajuste que já existe, em vez de criar outro.

    É o que faz dez técnicos reclamando da mesma coisa virarem um item com dez
    relatos — sem isso, o backlog encheria de duplicata e ninguém saberia qual
    ponto dói mais."""
    obs = db.query(models.TecnicoObservacao).filter(models.TecnicoObservacao.id == observacao_id).first()
    if not obs:
        raise HTTPException(status_code=404, detail="Observação não encontrada")
    ajuste = db.query(models.AtivoAjuste).filter(models.AtivoAjuste.id == payload.ajuste_id).first()
    if not ajuste:
        raise HTTPException(status_code=404, detail="Ajuste não encontrado")
    obs.ajuste_id = ajuste.id
    db.commit()
    tecnico = _get_tecnico_or_404(db, obs.tecnico_id)
    db.refresh(tecnico)
    return tecnico


@router.post("/tecnicos/limpar")
def limpar_base(payload: schemas.LimparBaseTecnicos, db: Session = Depends(get_db)):
    """Zera a base de técnicos — apaga todos os cadastros e o feedback junto.

    Serve pra recomeçar a importação do zero quando a planilha de origem estava
    errada. Não tem volta, então exige a palavra APAGAR: um DELETE disparado sem
    querer (ou um clique errado na tela) não passa daqui.
    """
    if (payload.confirmar or "").strip().upper() != "APAGAR":
        raise HTTPException(
            status_code=400,
            detail='Confirmação inválida — para zerar a base, mande {"confirmar": "APAGAR"}.',
        )
    tecnicos = db.query(models.Tecnico).count()
    observacoes = db.query(models.TecnicoObservacao).count()
    # as observações são apagadas explicitamente: no delete em massa quem cuidaria
    # disso seria o ON DELETE CASCADE do banco, e o SQLite local não o aplica por
    # padrão — sem isso, o histórico ficaria órfão em vez de sumir junto
    db.query(models.TecnicoObservacao).delete(synchronize_session=False)
    db.query(models.Tecnico).delete(synchronize_session=False)
    db.commit()
    return {"tecnicos_apagados": tecnicos, "observacoes_apagadas": observacoes}


@router.post("/tecnicos", response_model=schemas.TecnicoOut, status_code=201)
def create_tecnico(payload: schemas.TecnicoCreate, db: Session = Depends(get_db)):
    nome = (payload.nome or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome vazio.")
    papel = (payload.papel or "tecnico").strip()
    if papel not in PAPEIS:
        raise HTTPException(status_code=400, detail='Papel inválido — use "tecnico" ou "lider".')
    tecnico = models.Tecnico(
        nome=nome,
        telefone=_norm_telefone(payload.telefone),
        papel=papel,
        regional=(payload.regional or "").strip() or None,
        lider_nome=(payload.lider_nome or "").strip() or None,
        autor=(payload.autor or "").strip() or None,
        token=models.novo_token_tecnico(),
        status="a_contatar",
    )
    db.add(tecnico)
    db.commit()
    db.refresh(tecnico)
    return tecnico


@router.patch("/tecnicos/{tecnico_id}", response_model=schemas.TecnicoOut)
def update_tecnico(tecnico_id: int, payload: schemas.TecnicoUpdate, db: Session = Depends(get_db)):
    tecnico = _get_tecnico_or_404(db, tecnico_id)
    if payload.nome is not None:
        nome = payload.nome.strip()
        if not nome:
            raise HTTPException(status_code=400, detail="Nome vazio.")
        tecnico.nome = nome
    if payload.telefone is not None:
        tecnico.telefone = _norm_telefone(payload.telefone)
    if payload.papel is not None:
        if payload.papel not in PAPEIS:
            raise HTTPException(status_code=400, detail='Papel inválido — use "tecnico" ou "lider".')
        tecnico.papel = payload.papel
    if payload.regional is not None:
        tecnico.regional = payload.regional.strip() or None
    if payload.lider_nome is not None:
        tecnico.lider_nome = payload.lider_nome.strip() or None
    if payload.status is not None:
        if payload.status not in STATUSES:
            raise HTTPException(status_code=400, detail="Status inválido.")
        tecnico.status = payload.status
        agora = datetime.now(timezone.utc)
        if payload.status == "convidado" and tecnico.convidado_em is None:
            tecnico.convidado_em = agora
        elif payload.status == "instalado" and tecnico.instalado_em is None:
            tecnico.instalado_em = agora
        elif payload.status == "concluido" and tecnico.concluido_em is None:
            tecnico.concluido_em = agora
    db.commit()
    db.refresh(tecnico)
    return tecnico


@router.delete("/tecnicos/{tecnico_id}")
def delete_tecnico(tecnico_id: int, db: Session = Depends(get_db)):
    tecnico = _get_tecnico_or_404(db, tecnico_id)
    db.delete(tecnico)
    db.commit()
    return {"deleted": tecnico_id}


@router.get("/tecnicos/{tecnico_id}/mensagem", response_model=schemas.TecnicoMensagemOut)
def mensagem_tecnico(tecnico_id: int, request: Request, tipo: str = "convite", db: Session = Depends(get_db)):
    """Monta a mensagem pronta (texto + link do WhatsApp) pra esse técnico:
    `tipo=convite` (padrão) chama pra instalação, `tipo=feedback` pede o retorno
    depois do atendimento e leva o link do formulário. O link do WhatsApp só
    pré-preenche texto; o APK/manual vai por fora, na própria conversa, já que o
    wa.me não anexa arquivo."""
    tecnico = _get_tecnico_or_404(db, tecnico_id)
    mensagem = _mensagem_para(tecnico, tipo=tipo, base_url=str(request.base_url))
    wa_link = f"https://wa.me/{tecnico.telefone}?text={quote(mensagem)}"
    return schemas.TecnicoMensagemOut(
        tecnico_id=tecnico.id, telefone=tecnico.telefone, mensagem=mensagem, wa_link=wa_link,
    )


@router.post("/tecnicos/{tecnico_id}/observacoes", response_model=schemas.TecnicoOut, status_code=201)
def add_observacao(tecnico_id: int, payload: schemas.TecnicoObservacaoCreate, db: Session = Depends(get_db)):
    tecnico = _get_tecnico_or_404(db, tecnico_id)
    texto = (payload.texto or "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Observação vazia.")
    tipo = (payload.tipo or "").strip() or None
    if tipo is not None and tipo not in TIPOS_OBS:
        raise HTTPException(status_code=400, detail='Tipo inválido — use "positivo", "melhoria" ou "problema".')
    db.add(models.TecnicoObservacao(
        tecnico_id=tecnico.id,
        autor=(payload.autor or "").strip() or None,
        texto=texto,
        tipo=tipo,
        chamado=(payload.chamado or "").strip() or None,
        versao_app=(payload.versao_app or "").strip() or _versao_da_fase(tecnico),
    ))
    db.commit()
    db.refresh(tecnico)
    return tecnico


@router.post("/tecnicos/observacoes/{observacao_id}/virar-ajuste", response_model=schemas.TecnicoOut)
def virar_ajuste(observacao_id: int, payload: schemas.VirarAjuste, db: Session = Depends(get_db)):
    """Transforma o relato de um técnico em item do backlog da Gestão de Ativos.

    É a ponte que faltava entre o piloto e o time de dev: o que o técnico contou
    vira um item no formato que a LP já trabalha ("como está hoje / como deve
    ser"), e a observação guarda qual ajuste ela gerou — assim dá pra ver, no card
    dele, que aquele ponto já foi levado, e não levantar duas vezes a mesma coisa.
    """
    obs = db.query(models.TecnicoObservacao).filter(models.TecnicoObservacao.id == observacao_id).first()
    if not obs:
        raise HTTPException(status_code=404, detail="Observação não encontrada")
    if obs.ajuste_id:
        raise HTTPException(status_code=400, detail="Esse relato já virou um ajuste.")
    tecnico = _get_tecnico_or_404(db, obs.tecnico_id)

    titulo = (payload.titulo or "").strip()
    if not titulo:
        raise HTTPException(status_code=400, detail="Título vazio.")
    # relato de problema é bug; o resto entra como melhoria
    tipo = (payload.tipo or "").strip() or ("Bug" if obs.tipo == "problema" else "Melhoria")
    if tipo not in ativos.TIPOS:
        raise HTTPException(status_code=400, detail="Tipo inválido — use Bug ou Melhoria.")
    prioridade = (payload.prioridade or "Média").strip()
    if prioridade not in ativos.PRIORIDADES:
        prioridade = "Média"
    versao = ativos._norm_versao(payload.versao) if payload.versao else _versao_ativa(db)

    quando = obs.created_at.strftime("%d/%m/%Y") if obs.created_at else "durante o teste"
    origem = f"Relatado por {tecnico.nome} no teste do Track One ({quando})."

    ajuste = models.AtivoAjuste(
        versao=versao,
        numero=ativos._next_numero(db, versao),
        titulo=titulo,
        tipo=tipo,
        area=(payload.area or "Track One (app do técnico)").strip() or None,
        prioridade=prioridade,
        atual=(payload.atual if payload.atual is not None else obs.texto),
        esperado=payload.esperado or "",
        observacao=(payload.observacao if payload.observacao is not None else origem),
        status="levantado",
        autor=(payload.autor or "").strip() or None,
    )
    db.add(ajuste)
    db.flush()          # precisa do id do ajuste pra amarrar na observação
    obs.ajuste_id = ajuste.id
    db.commit()
    db.refresh(tecnico)
    return tecnico


@router.delete("/tecnicos/observacoes/{observacao_id}", response_model=schemas.TecnicoOut)
def delete_observacao(observacao_id: int, db: Session = Depends(get_db)):
    obs = db.query(models.TecnicoObservacao).filter(models.TecnicoObservacao.id == observacao_id).first()
    if not obs:
        raise HTTPException(status_code=404, detail="Observação não encontrada")
    tecnico = _get_tecnico_or_404(db, obs.tecnico_id)
    db.delete(obs)
    db.commit()
    db.refresh(tecnico)
    return tecnico


@pagina_router.get("/formulario/{token}", response_class=HTMLResponse)
def pagina_formulario(token: str, db: Session = Depends(get_db)):
    """A página que o técnico abre no celular. Link errado devolve uma página
    explicando, não um JSON de erro — quem abre isso não é desenvolvedor."""
    tecnico = _tecnico_por_token(db, token)
    if not tecnico:
        return HTMLResponse(formulario.pagina_invalida(), status_code=404)
    ja_respondeu = ""
    if tecnico.respondido_em:
        quando = tecnico.respondido_em
        if quando.tzinfo is None:
            quando = quando.replace(tzinfo=timezone.utc)
        ja_respondeu = quando.astimezone(BRT).strftime("%d/%m")
    return HTMLResponse(
        formulario.montar_html(tecnico.nome, tecnico.token, ja_respondeu),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


def _versao_ativa(db: Session) -> str:
    """A leva mais recente do backlog da Gestão de Ativos — é pra onde vai o item
    criado a partir de um relato, pra ele cair junto do que o time está tratando
    agora, e não numa rodada antiga já fechada."""
    versoes = [v for (v,) in db.query(models.AtivoAjuste.versao).distinct().all() if v]
    if not versoes:
        return "v2"

    def rank(v: str) -> int:
        digitos = "".join(c for c in v if c.isdigit())
        return int(digitos) if digitos else 0

    return max(versoes, key=rank)


def _tecnico_por_token(db: Session, token: str) -> Optional[models.Tecnico]:
    token = (token or "").strip()
    if not token:
        return None
    return db.query(models.Tecnico).filter(models.Tecnico.token == token).first()


@router.post("/formulario/{token}", response_model=schemas.TecnicoOut)
def responder_formulario(token: str, payload: schemas.FormularioResposta, db: Session = Depends(get_db)):
    """Recebe o formulário que o técnico preencheu no celular.

    Cada campo preenchido vira uma observação no card dele com o tipo certo, a
    nota e as etapas ficam no próprio técnico e o status vai pra "concluído" —
    é o retorno chegando na base sem ninguém digitar nada.
    """
    tecnico = _tecnico_por_token(db, token)
    if not tecnico:
        raise HTTPException(status_code=404, detail="Formulário não encontrado.")

    respostas = [
        ("positivo", (payload.positivo or "").strip()),
        ("melhoria", (payload.melhoria or "").strip()),
        ("problema", (payload.problema or "").strip()),
        (None, (payload.comentario or "").strip()),
    ]
    etapas = [e.strip() for e in (payload.etapas or []) if e and e.strip()]
    nota = payload.nota if payload.nota in (1, 2, 3, 4, 5) else None
    if not nota and not etapas and not any(texto for _, texto in respostas):
        raise HTTPException(status_code=400, detail="Responda pelo menos um campo.")

    chamado = (payload.chamado or "").strip() or None
    versao = _versao_da_fase(tecnico)
    for tipo, texto in respostas:
        if texto:
            db.add(models.TecnicoObservacao(
                tecnico_id=tecnico.id, autor=tecnico.nome, texto=texto, tipo=tipo,
                chamado=chamado, versao_app=versao,
            ))
    if nota:
        tecnico.nota = nota
    if etapas:
        tecnico.etapas_testadas = "|".join(etapas)
    tecnico.respondido_em = datetime.now(timezone.utc)
    # respondeu = testou; só não rebaixa quem já tinha sido marcado como concluído
    if tecnico.status != "concluido":
        tecnico.status = "concluido"
        if tecnico.concluido_em is None:
            tecnico.concluido_em = tecnico.respondido_em
    db.commit()
    db.refresh(tecnico)
    return tecnico


# colunas aceitas na planilha de importação — cada chave aceita algumas variações
# de nome (sem acento/maiúscula, já que _norm_txt normaliza os dois lados)
COLUNAS_IMPORTACAO = {
    "nome": {"nome", "tecnico", "nome do tecnico", "nome completo"},
    "telefone": {"telefone", "whatsapp", "celular", "fone", "numero"},
    "papel": {"papel", "funcao", "função", "cargo", "tipo"},
    "regional": {"regional", "cidade", "regiao", "praca"},
    "lider_nome": {"lider", "lider_nome", "lider direto", "responde a", "supervisor"},
}


@router.get("/tecnicos/exportar")
def exportar_piloto(fase_id: Optional[int] = None, db: Session = Depends(get_db)):
    """O piloto como planilha — pra levar pra reunião sem depender do link.

    Três abas: os técnicos com o andamento de cada um, o feedback linha a linha
    (com chamado, versão e o ajuste que gerou) e o resumo com os critérios.
    """
    fase = None
    if fase_id:
        fase = db.query(models.PilotoFase).filter(models.PilotoFase.id == fase_id).first()
        if not fase:
            raise HTTPException(status_code=404, detail="Fase não encontrada")

    q = db.query(models.Tecnico).options(
        joinedload(models.Tecnico.observacoes).joinedload(models.TecnicoObservacao.ajuste)
    )
    if fase_id:
        q = q.filter(models.Tecnico.fase_id == fase_id)
    tecnicos = q.order_by(models.Tecnico.nome).all()
    painel = painel_piloto(fase_id=fase_id, db=db)

    wb = Workbook()
    cabecalho_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    def escreve_cabecalho(ws, colunas, larguras):
        ws.append(colunas)
        for i in range(1, len(colunas) + 1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = cabecalho_fill
        for i, w in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Técnicos"
    escreve_cabecalho(
        ws,
        ["Nome", "Telefone", "Papel", "Regional", "Líder", "Situação", "Nota",
         "Etapas testadas", "Respondeu em", "Anotações"],
        [28, 18, 10, 18, 20, 16, 7, 46, 18, 11],
    )
    for t in tecnicos:
        ws.append([
            t.nome, t.telefone,
            "Líder" if t.papel == "lider" else "Técnico",
            t.regional or "", t.lider_nome or "",
            STATUS_LABELS.get(t.status, t.status),
            t.nota or "",
            (t.etapas_testadas or "").replace("|", " · "),
            t.respondido_em.astimezone(BRT).strftime("%d/%m/%Y %H:%M") if t.respondido_em else "",
            len(t.observacoes),
        ])

    ws2 = wb.create_sheet("Feedback")
    escreve_cabecalho(
        ws2,
        ["Técnico", "Tipo", "Relato", "Chamado", "Versão", "Virou ajuste", "Quando"],
        [26, 12, 70, 12, 10, 14, 18],
    )
    TIPO_LABEL = {"positivo": "Achou bom", "melhoria": "Melhoria", "problema": "Problema"}
    for t in tecnicos:
        for o in t.observacoes:
            ws2.append([
                t.nome, TIPO_LABEL.get(o.tipo, "Nota geral"), o.texto,
                o.chamado or "", o.versao_app or "", o.ajuste_ref or "",
                o.created_at.astimezone(BRT).strftime("%d/%m/%Y %H:%M") if o.created_at else "",
            ])
    for linha in ws2.iter_rows(min_row=2):
        linha[2].alignment = Alignment(wrap_text=True, vertical="top")

    ws3 = wb.create_sheet("Resumo")
    escreve_cabecalho(ws3, ["Indicador", "Atual", "Meta", "Situação"], [46, 12, 12, 16])
    ws3.append(["Técnicos na fase", painel["total"], "", ""])
    ws3.append(["Responderam o formulário", painel["responderam"], "", ""])
    ws3.append(["Nota média", painel["notas"]["media"] or "—", "", ""])
    ws3.append([])
    for c in painel["criterios"]:
        ws3.append([c["nome"], c["atual"], c["meta"], "OK" if c["ok"] else "Falta"])
    ws3.append([])
    ws3.append(["Cobertura do fluxo", "", "", ""])
    for c in painel["cobertura"]:
        ws3.append([f"  {c['etapa']}", c["total"], "", "OK" if c["ok"] else "Abaixo"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    apelido = "".join(ch for ch in (fase.nome if fase else "piloto") if ch.isalnum() or ch in " -_").strip()
    nome_arquivo = f"TrackOne_{apelido or 'piloto'}_{datetime.now(BRT).strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.get("/tecnicos/modelo")
def baixar_modelo_importacao():
    """Planilha modelo pra importar a base de técnicos de uma vez em /tecnicos/importar."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Técnicos"
    ws.append(["nome", "telefone", "papel", "regional", "lider"])
    ws.append(["João Silva", "(11) 99999-8888", "tecnico", "SP capital", "Carlos Souza"])
    ws.append(["Marcos Souza", "(11) 98888-7777", "lider", "Interior SP", ""])
    ws.column_dimensions[get_column_letter(1)].width = 26
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 12
    ws.column_dimensions[get_column_letter(4)].width = 20
    ws.column_dimensions[get_column_letter(5)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_tecnicos.xlsx"'},
    )


@router.post("/tecnicos/importar")
async def importar_tecnicos(
    file: UploadFile = File(...), autor: Optional[str] = Form(default=None), db: Session = Depends(get_db)
):
    """Sobe a base de técnicos de uma vez a partir de uma planilha .xlsx (colunas:
    nome, telefone e, opcionais, papel/regional/lider — ver /tecnicos/modelo).

    Quem já está na base (mesmo telefone) tem o cadastro atualizado em vez de ser
    rejeitado — reimportar a planilha nova é sincronizar, não duplicar; o que é
    progresso de QA (status, nota, etapas, feedback) nunca é tocado. Linha sem
    nome/telefone, telefone inválido ou telefone repetido dentro da própria
    planilha entra no relatório de rejeitados, sem travar as demais."""
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx.")
    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Não consegui abrir a planilha — confira o formato.")
    linhas = list(wb.active.iter_rows(values_only=True))
    if not linhas:
        raise HTTPException(status_code=400, detail="Planilha vazia.")

    cabecalho = [_norm_txt(c) for c in linhas[0]]
    idx = {}
    usadas = set()
    for campo, aliases in COLUNAS_IMPORTACAO.items():
        # 1ª passada: coluna com o nome exato (ex.: "Nome", "Telefone"). 2ª passada
        # (fallback): a base pode vir de outro sistema com coluna composta — ex.
        # "Endereço - Cidade" cai em "regional" por conter "cidade".
        achou = None
        for i, h in enumerate(cabecalho):
            if i not in usadas and h in aliases:
                achou = i
                break
        if achou is None:
            for i, h in enumerate(cabecalho):
                if i not in usadas and any(a in h for a in aliases):
                    achou = i
                    break
        if achou is not None:
            idx[campo] = achou
            usadas.add(achou)
    if "nome" not in idx or "telefone" not in idx:
        raise HTTPException(status_code=400, detail="A planilha precisa ter as colunas 'nome' e 'telefone'.")

    def valor(row, campo):
        i = idx.get(campo)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    # quem já está na base, pelo telefone — reimportar a planilha atualizada não
    # rejeita esse pessoal, atualiza o cadastro deles (ver abaixo)
    ja_na_base = {t.telefone: t for t in db.query(models.Tecnico).all()}
    vistos_no_arquivo = {}   # telefone -> linha em que apareceu primeiro

    criados = 0
    atualizados = 0
    em_branco = 0
    rejeitados = []
    for linha_num, row in enumerate(linhas[1:], start=2):
        nome = valor(row, "nome")
        telefone_bruto = valor(row, "telefone")
        if not nome and not telefone_bruto:
            em_branco += 1
            continue  # linha em branco — não é erro, só não tem nada nela

        def rejeita(motivo):
            rejeitados.append({
                "linha": linha_num, "nome": nome, "telefone": telefone_bruto, "motivo": motivo,
            })

        if not telefone_bruto:
            rejeita("sem telefone")
            continue
        if not nome:
            rejeita("sem nome")
            continue
        try:
            telefone = _norm_telefone(telefone_bruto)
        except HTTPException:
            rejeita("telefone inválido")
            continue
        if telefone in vistos_no_arquivo:
            rejeita(f"telefone repetido na planilha (já veio na linha {vistos_no_arquivo[telefone]})")
            continue
        vistos_no_arquivo[telefone] = linha_num

        papel = "lider" if _norm_txt(valor(row, "papel")) in ("lider", "lider de equipe", "supervisor") else "tecnico"
        regional = valor(row, "regional") or None
        lider_nome = valor(row, "lider_nome") or None

        existente = ja_na_base.get(telefone)
        if existente:
            # atualiza o cadastro, nunca o andamento do teste: status, nota,
            # etapas e feedback são progresso de QA e não vêm da planilha
            existente.nome = nome
            existente.papel = papel
            if regional:
                existente.regional = regional
            if lider_nome:
                existente.lider_nome = lider_nome
            if not existente.token:
                existente.token = models.novo_token_tecnico()
            atualizados += 1
            continue

        novo = models.Tecnico(
            nome=nome, telefone=telefone, papel=papel,
            regional=regional, lider_nome=lider_nome,
            autor=(autor or "").strip() or None,
            token=models.novo_token_tecnico(),
            status="a_contatar",
        )
        db.add(novo)
        ja_na_base[telefone] = novo
        criados += 1
    db.commit()

    # o resumo é o que responde "por que entraram menos do que a planilha tinha":
    # cada motivo com a sua contagem, em vez de uma lista de mil linhas soltas
    resumo = {}
    for r in rejeitados:
        chave = r["motivo"].split(" (já veio")[0]
        resumo[chave] = resumo.get(chave, 0) + 1
    return {
        "criados": criados,
        "atualizados": atualizados,
        "linhas_planilha": max(len(linhas) - 1, 0),
        "linhas_em_branco": em_branco,
        "rejeitados": len(rejeitados),
        "resumo": resumo,
        "erros": rejeitados,
    }
