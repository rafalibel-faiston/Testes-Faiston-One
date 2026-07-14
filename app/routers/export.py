import io
import re
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db

router = APIRouter(tags=["export"])

HEADERS = ["#", "Estágio", "Problema encontrado", "Ajuste solicitado", "Status"]

# pior status primeiro — o cabeçalho do estágio herda o pior status entre os itens dele,
# igual à planilha original (um estágio só é "Concluido" se tudo dentro dele passou)
STATUS_PRIORITY = ["Reprovado", "Bloqueado", "Não testado", "N/A", "Aprovado"]
STATUS_FILL = {
    "Aprovado": "D1FAE5",
    "Reprovado": "FEE2E2",
    "Bloqueado": "FEF3C7",
    "N/A": "E5E7EB",
    "Não testado": "F3F4F6",
}
STAGE_PREFIX_RE = re.compile(r"^\s*\d+\s*·\s*(.+)$")


def _stage_status(statuses):
    for s in STATUS_PRIORITY:
        if s in statuses:
            return s
    return "Não testado"


def _stage_label(estagio, estagio_num):
    """Tira o prefixo numérico ('01 · Criar OS' -> 'Criar OS') quando o estágio
    já tem número próprio — o número vai na coluna '#', igual à planilha original."""
    if estagio_num is not None:
        m = STAGE_PREFIX_RE.match(estagio or "")
        if m:
            return m.group(1)
    return estagio or ""


@router.get("/export")
def export_excel(fluxo: str = Query("C"), db: Session = Depends(get_db)):
    cases = (
        db.query(models.TestCase)
        .filter(models.TestCase.active.is_(True), models.TestCase.fluxo == fluxo)
        .order_by(models.TestCase.grupo, models.TestCase.estagio_num.nulls_last(), models.TestCase.code)
        .all()
    )

    groups = {}
    order = []
    for c in cases:
        key = ("num", c.estagio_num) if c.estagio_num is not None else ("label", c.grupo, c.estagio)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(c)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Fluxo {fluxo}"[:31]

    today = datetime.now().strftime("%d/%m/%Y")
    ws.merge_cells("A1:E1")
    ws["A1"] = f"STATUS DETALHADO DO FLUXO {fluxo} — Console de Teste (Faiston) · {today}"
    ws["A1"].font = Font(bold=True, size=13)

    ws.append([])
    ws.append(HEADERS)
    header_row = ws.max_row
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    if not cases:
        ws.append(["", "Nenhum caso de teste ativo neste fluxo.", "", "", ""])
    else:
        for key in order:
            items = groups[key]
            first = items[0]
            stage_num = first.estagio_num if first.estagio_num is not None else "-"
            stage_label = _stage_label(first.estagio, first.estagio_num)
            stage_status = _stage_status({c.status for c in items})

            ws.append([stage_num, stage_label, "", "", stage_status])
            r = ws.max_row
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=col_idx).font = Font(bold=True)
            fill = STATUS_FILL.get(stage_status)
            if fill:
                ws.cell(row=r, column=5).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")

            for c in items:
                ws.append(["-", c.frente, c.passos, c.resultado_esperado, c.status])
                r = ws.max_row
                fill = STATUS_FILL.get(c.status)
                if fill:
                    ws.cell(row=r, column=5).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {1: 6, 2: 24, 3: 46, 4: 46, 5: 15}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Fluxo{fluxo}_status_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
