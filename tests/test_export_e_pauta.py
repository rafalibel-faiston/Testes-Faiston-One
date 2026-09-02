"""Exportação do piloto e a entrada dele na pauta da reunião.

São as duas saídas do módulo pra fora da tela: a planilha que vai pra reunião e
a seção que a LP lê em /relatorio.
"""
import io

from openpyxl import load_workbook


def _com_feedback(client, tecnico):
    client.post(f"/api/formulario/{tecnico['token']}", json={
        "nota": 4,
        "etapas": ["Recebi o chamado no app", "Fechei a RAT pelo app"],
        "positivo": "rastreio ajudou",
        "problema": "notificação não chegou",
        "chamado": "221207",
    })


def test_exportacao_traz_tecnicos_feedback_e_resumo(client, tecnico):
    _com_feedback(client, tecnico)
    resp = client.get("/api/tecnicos/exportar")

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Técnicos", "Feedback", "Resumo"]

    linha = [c.value for c in wb["Técnicos"][2]]
    assert linha[0] == "Carlos Eduardo Lima"
    assert linha[6] == 4                                  # nota
    assert "Fechei a RAT pelo app" in linha[7]            # etapas testadas

    feedback = [[c.value for c in l] for l in wb["Feedback"].iter_rows(min_row=2)]
    tipos = {l[1] for l in feedback}
    assert tipos == {"Achou bom", "Problema"}
    assert all(l[3] == "221207" for l in feedback)        # chamado em cada relato

    resumo = [c.value for l in wb["Resumo"].iter_rows(min_row=2) for c in l[:1]]
    assert "Técnicos na fase" in resumo
    assert any("teste concluído" in str(v) for v in resumo)   # os critérios entram


def test_exportacao_de_uma_fase_leva_so_quem_esta_nela(client, tecnico):
    fora = client.post("/api/tecnicos", json={"nome": "Fora da fase", "telefone": "11977776666"}).json()
    fase = client.post("/api/piloto/fases", json={"nome": "Fase 1"}).json()
    client.post(f"/api/piloto/fases/{fase['id']}/tecnicos", json={"tecnico_ids": [tecnico["id"]]})

    wb = load_workbook(io.BytesIO(client.get(f"/api/tecnicos/exportar?fase_id={fase['id']}").content))
    nomes = [l[0].value for l in wb["Técnicos"].iter_rows(min_row=2)]
    assert nomes == [tecnico["nome"]]
    assert fora["nome"] not in nomes


def test_piloto_aparece_na_pauta_da_reuniao(client, tecnico):
    fase = client.post("/api/piloto/fases", json={"nome": "Fase 1 — SP capital", "versao_app": "1.2.0"}).json()
    client.post(f"/api/piloto/fases/{fase['id']}/tecnicos", json={"tecnico_ids": [tecnico["id"]]})
    _com_feedback(client, tecnico)

    pauta = client.get("/relatorio").text
    assert "Track One — piloto com os técnicos" in pauta
    assert "Fase 1 — SP capital" in pauta
    assert "1.2.0" in pauta
    # o que ainda não virou item de backlog é o que ninguém está tratando
    assert "notificação não chegou" in pauta


def test_relato_que_virou_ajuste_sai_da_lista_de_pendentes_da_pauta(client, tecnico):
    _com_feedback(client, tecnico)
    obs = [o for o in client.get("/api/tecnicos").json()[0]["observacoes"] if o["tipo"] == "problema"][0]

    assert "notificação não chegou" in client.get("/relatorio").text
    client.post(f"/api/tecnicos/observacoes/{obs['id']}/virar-ajuste", json={
        "titulo": "Notificação não dispara", "esperado": "push na entrega",
    })

    pauta = client.get("/relatorio").text
    # agora ele é acompanhado como ajuste, e não mais como ponto solto do piloto
    assert "Notificação não dispara" in pauta
    assert pauta.count("notificação não chegou") == 1     # só como texto do ajuste


def test_pauta_expoe_os_dados_do_piloto_pro_script(client, tecnico):
    """O script que gera a pauta como arquivo lê tudo pela API."""
    client.post("/api/piloto/fases", json={"nome": "Fase 1"})
    dados = client.get("/api/piloto/pauta").json()
    assert dados["fases"][0]["nome"] == "Fase 1"
    assert "relatos_sem_ajuste" in dados
